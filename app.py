import io
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from weekly_control import (
    DEFAULT_EXCLUSIONS,
    SupabaseRepository,
    daily_summary,
    normalize_text as normalize_weekly_text,
    read_detentions_excel,
    similarity,
    week_bounds,
    work_order_records,
)

try:
    from supabase import create_client
except ImportError:
    create_client = None

URL_CENTRO_APLICACIONES = (
    "https://centro-aplicaciones-area-planificacion.streamlit.app/"
)


st.set_page_config(
    page_title="Validación OT",
    page_icon="📋",
    layout="wide",
)

APP_DIR = Path(__file__).resolve().parent
LOGO_PATH = APP_DIR / "Finning-CAT.png"

OK_COLOR = "#FFC400"
OBS_COLOR = "#EF4444"


# ============================================================
# Estilos visuales
# ============================================================
st.markdown(
    """
    <style>
    .stApp {
        background: #f6f7f9;
    }

    .main-title {
        font-size: 34px;
        font-weight: 900;
        color: #111111;
        margin-bottom: 0;
    }

    .subtitle {
        font-size: 16px;
        color: #5b6470;
        margin-top: 0;
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #090909 0%, #1a1a1a 100%);
    }

    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] small {
        color: #ffffff;
    }

    section[data-testid="stSidebar"] div[data-testid="stImage"] {
        background: #ffffff;
        border-radius: 6px;
        padding: 4px;
        border: 1px solid #ffffff55;
    }

    .step-card {
        border: 1px solid #444444;
        border-radius: 10px;
        padding: 14px;
        margin-bottom: 14px;
        background: #141414;
        color: #ffffff;
    }

    .step-card b,
    .step-card small {
        color: #ffffff !important;
    }

    .step-number {
        display: inline-block;
        background: #ffc400;
        color: #111111 !important;
        border-radius: 50%;
        width: 28px;
        height: 28px;
        text-align: center;
        font-weight: 900;
        line-height: 28px;
        margin-right: 8px;
    }

    div[data-testid="stFileUploader"] {
        border: 1px dashed #ffc400;
        border-radius: 10px;
        padding: 8px;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] {
        background-color: #ffffff !important;
        border-radius: 10px !important;
        border: 1px dashed #ffc400 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] p {
        color: #111111 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] small {
        color: #5b6470 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] {
        color: #111111 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] p {
        color: #111111 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] small {
        color: #5b6470 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] button {
        background-color: #ffc400 !important;
        color: #111111 !important;
        border: 1px solid #d9a700 !important;
        font-weight: 800 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] button p,
    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] button span,
    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] button div {
        color: #111111 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] svg {
        color: #111111 !important;
        fill: #111111 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderFile"] {
        background-color: #ffffff !important;
        color: #111111 !important;
        border-radius: 6px;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderFile"] * {
        color: #111111 !important;
    }

    .kpi-card {
        background: #ffffff;
        border: 1px solid #e1e5ea;
        border-radius: 14px;
        padding: 18px 20px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
        min-height: 115px;
    }

    .kpi-title {
        font-size: 13px;
        color: #111111;
        font-weight: 800;
        text-transform: uppercase;
    }

    .kpi-value {
        font-size: 34px;
        color: #111111;
        font-weight: 900;
        line-height: 1.1;
    }

    .kpi-note {
        font-size: 13px;
        color: #657080;
    }

    .panel {
        background: #ffffff;
        border: 1px solid #e1e5ea;
        border-radius: 14px;
        padding: 18px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    }

    div[data-testid="stDownloadButton"] button,
    div.stButton > button {
        background: #ffc400 !important;
        color: #111111 !important;
        border: none !important;
        font-weight: 800 !important;
        border-radius: 8px !important;
    }

    div[data-testid="stDownloadButton"] button *,
    div.stButton > button * {
        color: #111111 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] > div > span {
        font-size: 0 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] > div > span::after {
        content: "Arrastra y suelta los archivos aquí";
        font-size: 16px !important;
        color: #111111 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] small {
        font-size: 0 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzoneInstructions"] small::after {
        content: "Límite 1 GB por archivo · XLSX, XLSM o ZIP";
        font-size: 13px !important;
        color: #5b6470 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] button {
        font-size: 0 !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stFileUploaderDropzone"] button::after {
        content: "Seleccionar archivos";
        font-size: 15px !important;
        color: #111111 !important;
    }

    /* ========================================================
       BOTÓN NATIVO: CENTRO DE APLICACIONES
    ======================================================== */
    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] {
        margin-top: 4px !important;
        margin-bottom: 18px !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] a {
        position: relative !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
        width: 100% !important;
        min-height: 50px !important;
        padding: 6px 8px 6px 58px !important;
        background: transparent !important;
        border: none !important;
        border-radius: 0 !important;
        color: #ffffff !important;
        text-decoration: none !important;
        box-shadow: none !important;
        transition: color 0.2s ease !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] a:hover {
        background: transparent !important;
        color: #ffc400 !important;
        transform: none !important;
    }

    /* Ícono de cuatro cuadros */
    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] a::before {
        content: "";
        position: absolute;
        left: 8px;
        top: 10px;
        width: 13px;
        height: 13px;
        border-radius: 3px;
        background: #ffffff;
        box-shadow:
            19px 0 0 #ffffff,
            0 19px 0 #ffffff,
            19px 19px 0 #ffffff;
    }

    /* Línea vertical amarilla */
    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] a::after {
        content: "";
        position: absolute;
        left: 47px;
        top: 6px;
        width: 3px;
        height: 38px;
        background: #ffc400;
        border-radius: 2px;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] a p {
        color: inherit !important;
        font-size: 17px !important;
        font-weight: 800 !important;
        line-height: 1.15 !important;
        margin: 0 !important;
        padding: 0 0 0 10px !important;
        border: none !important;
        white-space: nowrap !important;
    }

    section[data-testid="stSidebar"]
    div[data-testid="stLinkButton"] a svg {
        display: none !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# Configuración del formato OT
# ============================================================
BLANK_TOKENS = {
    "",
    "-",
    "--",
    "N/A",
    "NA",
    "NAN",
    "NONE",
    "NULL",
    "S/I",
    "SIN INFORMACION",
    "SIN INFORMACIÓN",
}

INVISIBLE_CHARS = ("\xa0", "\u200b", "\u200c", "\u200d", "\ufeff", "\u3164")

PRIORITY_CRITICAL = "Crítico"
PRIORITY_STANDARD = "Estándar"

CRITICAL_FIELDS = {
    "Código trabajo",
    "Código síntoma",
    "Código causa",
    "Firma jefe turno (nombre + RUT)",
    "Firma técnico responsable (nombre + RUT)",
}

VALIDATED_FIELDS = [
    "Horómetro",
    "Motivo detención del equipo",
    "Descripción del síntoma",
    "Código trabajo",
    "Código síntoma",
    "Código causa",
    "Descripción de actividades",
    "Firma jefe turno (nombre + RUT)",
    "Firma técnico responsable (nombre + RUT)",
]

INVALID_CAUSE_CODES = {6.6, 7.1}

WORK_ROWS = list(range(42, 94, 4))

WORK_FIELDS: List[Tuple[str, str]] = [
    ("B", "Hora inicio"),
    ("F", "Hora término"),
    ("J", "N° orden servicio"),
    ("Q", "Código componente SMCS"),
    ("T", "Código modificador"),
    ("W", "Código trabajo"),
    ("Z", "Descripción del síntoma"),
    ("AO", "Código síntoma"),
    ("AR", "Descripción de la causa"),
    ("BH", "Código causa"),
    ("BO", "Tipo tarea"),
    ("BV", "Tarea principal"),
]

ESSENTIAL_WORK_FIELDS: List[Tuple[str, str]] = [
    ("W", "Código trabajo"),
    ("Z", "Descripción del síntoma"),
    ("AO", "Código síntoma"),
    ("BH", "Código causa"),
]

SUMMARY_COLUMNS = [
    "Archivo",
    "Equipo",
    "Orden",
    "Turno",
    "Campos faltantes",
    "Faltantes críticos",
    "Estado",
    "Campos con observación",
]

# Orden ya no forma parte del detalle.
DETAIL_COLUMNS = [
    "Archivo",
    "Equipo",
    "Técnico responsable",
    "Jefe de turno",
    "Sección",
    "Campo faltante",
    "Prioridad",
    "Fila OT",
    "Celda/Rango",
    "Observación",
]


# ============================================================
# Utilidades de lectura y validación
# ============================================================
def normalize_value(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)

    for char in INVISIBLE_CHARS:
        text = text.replace(char, " ")

    text = " ".join(text.split()).strip()

    if text.upper() in BLANK_TOKENS:
        return ""

    return text


def is_filled(value: Any) -> bool:
    return normalize_value(value) != ""


def is_zero_numeric(value: Any) -> bool:
    text = normalize_value(value).replace(",", ".")

    if not text:
        return False

    try:
        return abs(float(text)) < 1e-9
    except (TypeError, ValueError):
        return False


def is_invalid_horometer_text(value: Any) -> bool:
    if value is None:
        return False

    text = str(value)

    for char in INVISIBLE_CHARS:
        text = text.replace(char, " ")

    text = " ".join(text.split()).strip().upper()

    return text in {
        "SIN HOROMETRO",
        "SIN HORÓMETRO",
        "SIN INFORMACION",
        "SIN INFORMACIÓN",
    }


def is_invalid_cause_code(value: Any) -> bool:
    text = normalize_value(value).replace(",", ".")

    if not text:
        return False

    try:
        numeric_value = float(text)
    except (TypeError, ValueError):
        return False

    return any(
        abs(numeric_value - invalid_code) < 1e-9
        for invalid_code in INVALID_CAUSE_CODES
    )


def summarize_observations(missing: Sequence[Dict[str, str]]) -> str:
    if not missing:
        return ""

    summaries: List[str] = []
    seen = set()

    for item in missing:
        field = item.get("Campo faltante", "Campo")
        observation = item.get("Observación", "")

        if observation.startswith("Valor inválido"):
            label = f"{field} ({observation})"
        elif observation.startswith("Falta "):
            label = f"{field} ({observation.lower()})"
        else:
            label = field

        if label not in seen:
            summaries.append(label)
            seen.add(label)

    return "; ".join(summaries)


def get_value(ws, cell: str) -> Any:
    try:
        return ws[cell].value
    except Exception:
        return None


def first_filled(ws, cells: Sequence[str]) -> str:
    for cell in cells:
        value = get_value(ws, cell)

        if is_filled(value):
            return normalize_value(value)

    return ""


def field_priority(field_name: str) -> str:
    return PRIORITY_CRITICAL if field_name in CRITICAL_FIELDS else PRIORITY_STANDARD


def add_missing(
    missing: List[Dict[str, str]],
    section: str,
    field: str,
    cell_range: str,
    observation: str = "Campo requerido sin información",
    row_label: str = "",
    priority: Optional[str] = None,
) -> None:
    missing.append(
        {
            "Sección": section,
            "Campo faltante": field,
            "Prioridad": priority or field_priority(field),
            "Fila OT": row_label,
            "Celda/Rango": cell_range,
            "Observación": observation,
        }
    )


def validate_single_cells(
    ws,
    missing: List[Dict[str, str]],
    section: str,
    fields: Iterable[Tuple[str, str]],
) -> None:
    for field, cell in fields:
        if not is_filled(get_value(ws, cell)):
            add_missing(missing, section, field, cell)


def active_rows(
    ws,
    rows: Sequence[int],
    columns: Sequence[str],
) -> List[int]:
    return [
        row
        for row in rows
        if any(
            is_filled(get_value(ws, f"{column}{row}"))
            for column in columns
        )
    ]


def validate_required_table(
    ws,
    missing: List[Dict[str, str]],
    section: str,
    rows: Sequence[int],
    required_fields: Sequence[Tuple[str, str]],
    trigger_columns: Optional[Sequence[str]] = None,
    require_one_row: bool = True,
    row_name: str = "Registro",
) -> List[int]:
    trigger_columns = list(
        trigger_columns or [column for column, _ in required_fields]
    )

    rows_used = active_rows(ws, rows, trigger_columns)
    rows_to_validate = rows_used

    if require_one_row and not rows_used:
        rows_to_validate = [rows[0]]

    for sequential_number, row in enumerate(rows_to_validate, start=1):
        row_label = f"{row_name} {sequential_number} (fila {row})"

        for column, field in required_fields:
            cell = f"{column}{row}"

            if not is_filled(get_value(ws, cell)):
                add_missing(
                    missing,
                    section,
                    field,
                    cell,
                    row_label=row_label,
                )

    return rows_used


def validate_signature(
    ws,
    missing: List[Dict[str, str]],
    field: str,
    name_cell: str,
    id_cell: str,
) -> None:
    missing_parts: List[str] = []

    if not is_filled(get_value(ws, name_cell)):
        missing_parts.append("nombre")

    if not is_filled(get_value(ws, id_cell)):
        missing_parts.append("RUT")

    if missing_parts:
        add_missing(
            missing,
            "Firmas y validación",
            field,
            f"{name_cell}/{id_cell}",
            f"Falta {' y '.join(missing_parts)} en el recuadro de firma",
            priority=PRIORITY_CRITICAL,
        )


def get_ot_worksheet(workbook):
    preferred_names = [
        "OT FORMATO IMPRIMIR",
        "OT",
        "ORDEN DE TRABAJO",
    ]

    normalized_names = {
        name.strip().upper(): name
        for name in workbook.sheetnames
    }

    for preferred_name in preferred_names:
        if preferred_name in normalized_names:
            return workbook[normalized_names[preferred_name]]

    return workbook[workbook.sheetnames[0]]


def validate_work_order(
    file_bytes: bytes,
    filename: str,
) -> Dict[str, Any]:
    workbook = load_workbook(
        io.BytesIO(file_bytes),
        data_only=True,
        read_only=False,
        keep_vba=filename.lower().endswith(".xlsm"),
    )

    ws = get_ot_worksheet(workbook)
    missing: List[Dict[str, str]] = []

    # 1) Horómetro
    horometer_cell = "G13"
    horometer_value = get_value(ws, horometer_cell)

    if (
        not is_filled(horometer_value)
        or is_zero_numeric(horometer_value)
        or is_invalid_horometer_text(horometer_value)
    ):
        add_missing(
            missing,
            "Información general",
            "Horómetro",
            horometer_cell,
            "Falta información de horómetro",
        )

    # 2) Motivo de detención
    validate_single_cells(
        ws,
        missing,
        "Motivo de detención",
        [("Motivo detención del equipo", "AB25")],
    )

    # 3) Información del trabajo
    work_rows_used = validate_required_table(
        ws,
        missing,
        "Información del trabajo",
        WORK_ROWS,
        ESSENTIAL_WORK_FIELDS,
        trigger_columns=[column for column, _ in WORK_FIELDS],
        require_one_row=True,
        row_name="Actividad",
    )

    rows_to_check = work_rows_used if work_rows_used else [WORK_ROWS[0]]

    for sequential_number, row in enumerate(rows_to_check, start=1):
        cell = f"BH{row}"
        value = get_value(ws, cell)

        if is_invalid_cause_code(value):
            add_missing(
                missing,
                "Información del trabajo",
                "Código causa",
                cell,
                (
                    f"Valor inválido {normalize_value(value)}: "
                    "los códigos 6.6 y 7.1 corresponden "
                    "a la categoría 'Otros'"
                ),
                row_label=f"Actividad {sequential_number} (fila {row})",
                priority=PRIORITY_CRITICAL,
            )

    # 4) Descripción de actividades
    activity_description_cells = [
        f"B{row}"
        for row in range(98, 125, 3)
    ]

    has_activity_description = any(
        is_filled(get_value(ws, cell))
        for cell in activity_description_cells
    )

    if not has_activity_description:
        add_missing(
            missing,
            "Descripción de actividades",
            "Descripción de actividades",
            "B98:BZ124",
        )

    # 5) Firmas
    validate_signature(
        ws,
        missing,
        "Firma jefe turno (nombre + RUT)",
        "C238",
        "C244",
    )

    validate_signature(
        ws,
        missing,
        "Firma técnico responsable (nombre + RUT)",
        "BD239",
        "BD243",
    )

    # Datos generales
    equipo = first_filled(ws, ["G7"])
    orden = first_filled(ws, ["G25", "J42", "Z145"])
    turno = first_filled(ws, ["G19"])

    # Datos adicionales utilizados por el Control OT Semanal.
    jefe_turno = first_filled(ws, ["C238"])
    tecnico_responsable = first_filled(ws, ["BD239"])
    motivo_detencion = normalize_value(get_value(ws, "AB25"))

    descripcion_partes: List[str] = []
    if motivo_detencion:
        descripcion_partes.append(motivo_detencion)

    for row in (work_rows_used if work_rows_used else WORK_ROWS):
        descripcion = normalize_value(get_value(ws, f"Z{row}"))
        if descripcion and descripcion not in descripcion_partes:
            descripcion_partes.append(descripcion)

    for cell in activity_description_cells:
        descripcion = normalize_value(get_value(ws, cell))
        if descripcion and descripcion not in descripcion_partes:
            descripcion_partes.append(descripcion)

    descripcion_ot = " | ".join(descripcion_partes)

    critical_missing = sum(
        1
        for item in missing
        if item["Prioridad"] == PRIORITY_CRITICAL
    )

    return {
        "Archivo": filename,
        "Equipo": equipo,
        "Orden": orden,
        "Turno": turno,
        "Técnico responsable": tecnico_responsable,
        "Jefe de turno": jefe_turno,
        "Motivo detención": motivo_detencion,
        "Descripción OT": descripcion_ot,
        "Campos faltantes": len(missing),
        "Faltantes críticos": critical_missing,
        "Estado": "Completa" if not missing else "Con observaciones",
        "Campos con observación": summarize_observations(missing),
        "missing": missing,
    }


# ============================================================
# Lectura de archivos subidos
# ============================================================
def read_uploaded_files(uploaded_files) -> List[Tuple[str, bytes]]:
    files: List[Tuple[str, bytes]] = []

    for uploaded in uploaded_files:
        name = uploaded.name
        data = uploaded.getvalue()
        lower_name = name.lower()

        if lower_name.endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(data)) as zip_file:
                    for info in zip_file.infolist():
                        internal_name = info.filename.replace("\\", "/")
                        file_name = internal_name.split("/")[-1]

                        is_excel = internal_name.lower().endswith(
                            (".xlsx", ".xlsm")
                        )

                        is_valid = (
                            is_excel
                            and not internal_name.startswith("__MACOSX")
                            and not file_name.startswith("~$")
                        )

                        if is_valid:
                            files.append((file_name, zip_file.read(info)))

            except zipfile.BadZipFile:
                continue

        elif (
            lower_name.endswith((".xlsx", ".xlsm"))
            and not name.startswith("~$")
        ):
            files.append((name, data))

    return files


# ============================================================
# Construcción de tablas
# ============================================================
def create_result_dataframes(
    results: List[Dict[str, Any]],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    summary_records: List[Dict[str, Any]] = []
    detail_records: List[Dict[str, Any]] = []

    for result in results:
        summary_records.append(
            {
                key: result.get(key, "")
                for key in SUMMARY_COLUMNS
            }
        )

        for item in result.get("missing", []):
            detail_records.append(
                {
                    "Archivo": result.get("Archivo", ""),
                    "Equipo": result.get("Equipo", ""),
                    "Técnico responsable": result.get(
                        "Técnico responsable",
                        "",
                    ),
                    "Jefe de turno": result.get(
                        "Jefe de turno",
                        "",
                    ),
                    **item,
                }
            )

    summary_df = pd.DataFrame(
        summary_records,
        columns=SUMMARY_COLUMNS,
    )

    detail_df = pd.DataFrame(
        detail_records,
        columns=DETAIL_COLUMNS,
    )

    return summary_df, detail_df


def build_field_summary(
    detail_df: pd.DataFrame,
    total_orders: int,
) -> pd.DataFrame:
    columns = [
        "Prioridad",
        "Campo faltante",
        "Cantidad faltante",
        "OT afectadas",
        "% OT afectadas",
    ]

    if detail_df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        detail_df.groupby(
            ["Prioridad", "Campo faltante"],
            as_index=False,
        )
        .agg(
            **{
                "Cantidad faltante": ("Campo faltante", "size"),
                "OT afectadas": ("Archivo", "nunique"),
            }
        )
    )

    summary["% OT afectadas"] = (
        summary["OT afectadas"]
        / max(total_orders, 1)
        * 100
    ).round(1)

    priority_order = {
        PRIORITY_CRITICAL: 0,
        PRIORITY_STANDARD: 1,
    }

    summary["_priority_order"] = (
        summary["Prioridad"]
        .map(priority_order)
        .fillna(9)
    )

    summary = (
        summary.sort_values(
            [
                "_priority_order",
                "Cantidad faltante",
                "Campo faltante",
            ],
            ascending=[True, False, True],
        )
        .drop(columns="_priority_order")
    )

    return summary[columns]


def build_compliance_summary(
    detail_df: pd.DataFrame,
    total_orders: int,
) -> pd.DataFrame:
    columns = [
        "Campo",
        "OT conformes",
        "OT con observación",
        "% Cumplimiento",
        "% Con observación",
    ]

    if total_orders <= 0:
        return pd.DataFrame(columns=columns)

    affected_by_field: Dict[str, int] = {}

    if not detail_df.empty:
        affected_by_field = (
            detail_df.groupby("Campo faltante")["Archivo"]
            .nunique()
            .astype(int)
            .to_dict()
        )

    records: List[Dict[str, Any]] = []

    for field in VALIDATED_FIELDS:
        affected = min(
            affected_by_field.get(field, 0),
            total_orders,
        )

        conforming = total_orders - affected

        records.append(
            {
                "Campo": field,
                "OT conformes": conforming,
                "OT con observación": affected,
                "% Cumplimiento": round(
                    conforming / total_orders * 100,
                    1,
                ),
                "% Con observación": round(
                    affected / total_orders * 100,
                    1,
                ),
            }
        )

    return pd.DataFrame(records, columns=columns)


def build_section_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Sección",
        "Campos faltantes",
        "% del total",
    ]

    if detail_df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        detail_df.groupby("Sección", as_index=False)
        .size()
        .rename(columns={"size": "Campos faltantes"})
    )

    total_missing = int(summary["Campos faltantes"].sum())

    summary["% del total"] = (
        summary["Campos faltantes"]
        / total_missing
        * 100
    ).round(1)

    return summary.sort_values(
        "Campos faltantes",
        ascending=False,
    )[columns]


# ============================================================
# Reporte Excel
# ============================================================
def build_excel_report(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    field_summary_df: pd.DataFrame,
    section_summary_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    total_orders = len(summary_df)
    orders_with_obs = (
        int((summary_df["Estado"] == "Con observaciones").sum())
        if total_orders
        else 0
    )
    complete_orders = (
        int((summary_df["Estado"] == "Completa").sum())
        if total_orders
        else 0
    )
    total_missing = (
        int(summary_df["Campos faltantes"].sum())
        if total_orders
        else 0
    )
    critical_missing = (
        int(summary_df["Faltantes críticos"].sum())
        if total_orders
        else 0
    )

    compliance_summary_df = build_compliance_summary(
        detail_df,
        total_orders,
    )

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Resumen OT",
        )
        compliance_summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Cumplimiento por campo",
        )
        field_summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Resumen por campo",
        )
        section_summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Resumen secciones",
        )
        detail_df.to_excel(
            writer,
            index=False,
            sheet_name="Detalle faltantes",
        )

        workbook = writer.book

        header_fmt = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#FFC400",
                "border": 1,
                "align": "center",
                "valign": "vcenter",
            }
        )
        body_fmt = workbook.add_format(
            {
                "border": 1,
                "valign": "top",
            }
        )
        bad_fmt = workbook.add_format(
            {
                "bg_color": "#FFECEC",
                "font_color": "#C00000",
                "border": 1,
            }
        )
        critical_fmt = workbook.add_format(
            {
                "bg_color": "#FECACA",
                "font_color": "#991B1B",
                "border": 1,
                "bold": True,
            }
        )
        ok_fmt = workbook.add_format(
            {
                "bg_color": "#FFC400",
                "font_color": "#111111",
                "border": 1,
            }
        )
        title_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 18,
                "bg_color": "#111111",
                "font_color": "#FFFFFF",
                "align": "center",
            }
        )
        kpi_title_fmt = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#FFC400",
                "border": 1,
                "align": "center",
            }
        )
        kpi_value_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 20,
                "border": 1,
                "align": "center",
            }
        )

        sheets = [
            ("Resumen OT", summary_df),
            ("Cumplimiento por campo", compliance_summary_df),
            ("Resumen por campo", field_summary_df),
            ("Resumen secciones", section_summary_df),
            ("Detalle faltantes", detail_df),
        ]

        for sheet_name, dataframe in sheets:
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes(1, 0)

            if len(dataframe.columns) > 0:
                worksheet.autofilter(
                    0,
                    0,
                    max(len(dataframe), 1),
                    len(dataframe.columns) - 1,
                )

            for col_num, col_name in enumerate(dataframe.columns):
                worksheet.write(0, col_num, col_name, header_fmt)

                width = min(
                    max(len(str(col_name)) + 4, 14),
                    45,
                )

                if not dataframe.empty:
                    max_content = (
                        dataframe[col_name]
                        .astype(str)
                        .map(len)
                        .max()
                    )
                    width = min(
                        max(width, int(max_content) + 3),
                        55,
                    )

                worksheet.set_column(
                    col_num,
                    col_num,
                    width,
                    body_fmt,
                )

        if not summary_df.empty:
            ws_summary = writer.sheets["Resumen OT"]
            estado_col = summary_df.columns.get_loc("Estado")
            critical_col = summary_df.columns.get_loc(
                "Faltantes críticos"
            )

            ws_summary.conditional_format(
                1,
                estado_col,
                len(summary_df),
                estado_col,
                {
                    "type": "text",
                    "criteria": "containing",
                    "value": "Con observaciones",
                    "format": bad_fmt,
                },
            )

            ws_summary.conditional_format(
                1,
                estado_col,
                len(summary_df),
                estado_col,
                {
                    "type": "text",
                    "criteria": "containing",
                    "value": "Completa",
                    "format": ok_fmt,
                },
            )

            ws_summary.conditional_format(
                1,
                critical_col,
                len(summary_df),
                critical_col,
                {
                    "type": "cell",
                    "criteria": ">",
                    "value": 0,
                    "format": critical_fmt,
                },
            )

        if not detail_df.empty:
            ws_detail = writer.sheets["Detalle faltantes"]
            priority_col = detail_df.columns.get_loc("Prioridad")

            ws_detail.conditional_format(
                1,
                priority_col,
                len(detail_df),
                priority_col,
                {
                    "type": "text",
                    "criteria": "containing",
                    "value": PRIORITY_CRITICAL,
                    "format": critical_fmt,
                },
            )

        dashboard = workbook.add_worksheet("Dashboard")
        writer.sheets["Dashboard"] = dashboard

        dashboard.hide_gridlines(2)
        dashboard.set_column("A:A", 3)
        dashboard.set_column("B:I", 23)
        dashboard.merge_range(
            "B2:I3",
            "REPORTE EJECUTIVO DE VALIDACIÓN OT",
            title_fmt,
        )

        kpis = [
            ("Órdenes procesadas", total_orders),
            ("Órdenes con observaciones", orders_with_obs),
            ("Hallazgos detectados", total_missing),
            ("Hallazgos críticos", critical_missing),
            ("Órdenes completas", complete_orders),
        ]

        for index, (title, value) in enumerate(kpis):
            col = 1 + index
            dashboard.write(4, col, title, kpi_title_fmt)
            dashboard.write(5, col, value, kpi_value_fmt)

        if not compliance_summary_df.empty:
            chart = workbook.add_chart(
                {
                    "type": "bar",
                    "subtype": "stacked",
                }
            )

            last_row = len(compliance_summary_df)

            chart.add_series(
                {
                    "name": "Conforme",
                    "categories": [
                        "Cumplimiento por campo",
                        1,
                        0,
                        last_row,
                        0,
                    ],
                    "values": [
                        "Cumplimiento por campo",
                        1,
                        3,
                        last_row,
                        3,
                    ],
                    "fill": {"color": "#FFC400"},
                    "border": {"color": "#D6A400"},
                    "data_labels": {
                        "value": True,
                        "num_format": r"0.0\%",
                    },
                }
            )

            chart.add_series(
                {
                    "name": "Con observación",
                    "categories": [
                        "Cumplimiento por campo",
                        1,
                        0,
                        last_row,
                        0,
                    ],
                    "values": [
                        "Cumplimiento por campo",
                        1,
                        4,
                        last_row,
                        4,
                    ],
                    "fill": {"color": "#EF4444"},
                    "border": {"color": "#B91C1C"},
                    "data_labels": {
                        "value": True,
                        "num_format": r"0.0\%",
                    },
                }
            )

            chart.set_title({"name": "Cumplimiento por campo"})
            chart.set_legend({"position": "bottom"})
            chart.set_x_axis(
                {
                    "name": "Porcentaje de OT",
                    "min": 0,
                    "max": 100,
                    "major_unit": 20,
                }
            )
            chart.set_y_axis({"reverse": True})
            chart.set_style(10)

            dashboard.insert_chart(
                "B9",
                chart,
                {
                    "x_scale": 1.5,
                    "y_scale": 1.45,
                },
            )

        status_data_row = 32

        dashboard.write(
            status_data_row,
            1,
            "Estado",
            header_fmt,
        )
        dashboard.write(
            status_data_row,
            2,
            "Cantidad",
            header_fmt,
        )
        dashboard.write(
            status_data_row + 1,
            1,
            "Completas",
            body_fmt,
        )
        dashboard.write(
            status_data_row + 1,
            2,
            complete_orders,
            body_fmt,
        )
        dashboard.write(
            status_data_row + 2,
            1,
            "Con observaciones",
            body_fmt,
        )
        dashboard.write(
            status_data_row + 2,
            2,
            orders_with_obs,
            body_fmt,
        )

        if total_orders:
            donut = workbook.add_chart({"type": "doughnut"})

            donut.add_series(
                {
                    "name": "Estado de órdenes",
                    "categories": [
                        "Dashboard",
                        status_data_row + 1,
                        1,
                        status_data_row + 2,
                        1,
                    ],
                    "values": [
                        "Dashboard",
                        status_data_row + 1,
                        2,
                        status_data_row + 2,
                        2,
                    ],
                    "points": [
                        {"fill": {"color": "#FFC400"}},
                        {"fill": {"color": "#EF4444"}},
                    ],
                    "data_labels": {
                        "percentage": True,
                        "category": True,
                    },
                }
            )

            donut.set_title({"name": "Estado de órdenes"})
            donut.set_hole_size(55)
            donut.set_legend({"position": "bottom"})

            dashboard.insert_chart(
                "K9",
                donut,
                {
                    "x_scale": 1.1,
                    "y_scale": 1.2,
                },
            )

    return output.getvalue()


# ============================================================
# Reporte PDF
# ============================================================
def build_pdf_report(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    field_summary_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )

    styles = getSampleStyleSheet()

    story = [
        Paragraph(
            "Reporte Ejecutivo de Validación de Órdenes de Trabajo",
            styles["Title"],
        ),
        Paragraph(
            f"Fecha reporte: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 0.3 * cm),
    ]

    total_orders = len(summary_df)
    orders_with_obs = (
        int((summary_df["Estado"] == "Con observaciones").sum())
        if total_orders
        else 0
    )
    complete_orders = (
        int((summary_df["Estado"] == "Completa").sum())
        if total_orders
        else 0
    )
    total_missing = (
        int(summary_df["Campos faltantes"].sum())
        if total_orders
        else 0
    )
    critical_missing = (
        int(summary_df["Faltantes críticos"].sum())
        if total_orders
        else 0
    )

    kpi_data = [
        [
            "Órdenes procesadas",
            "Órdenes con observaciones",
            "Hallazgos detectados",
            "Hallazgos críticos",
            "Órdenes completas",
        ],
        [
            total_orders,
            orders_with_obs,
            total_missing,
            critical_missing,
            complete_orders,
        ],
    ]

    kpi_table = Table(
        kpi_data,
        colWidths=[5.0 * cm] * 5,
        rowHeights=[1.0 * cm, 0.9 * cm],
    )

    kpi_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#FFC400"),
                ),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, 1), 18),
            ]
        )
    )

    story.extend(
        [
            kpi_table,
            Spacer(1, 0.4 * cm),
        ]
    )

    story.append(
        Paragraph(
            "Resumen por OT",
            styles["Heading2"],
        )
    )

    pdf_summary_columns = [
        "Equipo",
        "Orden",
        "Estado",
        "Campos faltantes",
    ]

    pdf_summary = (
        summary_df[pdf_summary_columns]
        .head(20)
        .copy()
    )

    pdf_summary_data = (
        [pdf_summary.columns.tolist()]
        + pdf_summary.astype(str).values.tolist()
    )

    pdf_summary_table = Table(
        pdf_summary_data,
        repeatRows=1,
        colWidths=[
            4.0 * cm,
            4.0 * cm,
            5.0 * cm,
            4.0 * cm,
        ],
    )

    pdf_summary_style = [
        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            colors.HexColor("#111111"),
        ),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
    ]

    for row_number, estado in enumerate(
        pdf_summary["Estado"].tolist(),
        start=1,
    ):
        if estado == "Completa":
            pdf_summary_style.append(
                (
                    "BACKGROUND",
                    (0, row_number),
                    (-1, row_number),
                    colors.HexColor("#FFC400"),
                )
            )
            pdf_summary_style.append(
                (
                    "TEXTCOLOR",
                    (0, row_number),
                    (-1, row_number),
                    colors.black,
                )
            )
        else:
            pdf_summary_style.append(
                (
                    "BACKGROUND",
                    (0, row_number),
                    (-1, row_number),
                    colors.HexColor("#FFECEC"),
                )
            )
            pdf_summary_style.append(
                (
                    "TEXTCOLOR",
                    (0, row_number),
                    (-1, row_number),
                    colors.HexColor("#991B1B"),
                )
            )

    pdf_summary_table.setStyle(
        TableStyle(pdf_summary_style)
    )

    story.extend(
        [
            pdf_summary_table,
            Spacer(1, 0.4 * cm),
        ]
    )

    story.append(
        Paragraph(
            "Campos faltantes por campo",
            styles["Heading2"],
        )
    )

    field_show = field_summary_df.head(15).copy()
    field_table_data = (
        [field_show.columns.tolist()]
        + field_show.astype(str).values.tolist()
    )

    field_table = Table(
        field_table_data,
        repeatRows=1,
    )

    field_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#FFC400"),
                ),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                (
                    "BACKGROUND",
                    (0, 1),
                    (0, -1),
                    colors.HexColor("#FFF7D6"),
                ),
            ]
        )
    )

    story.extend(
        [
            field_table,
            Spacer(1, 0.4 * cm),
        ]
    )

    story.append(
        Paragraph(
            "Detalle de órdenes con observaciones",
            styles["Heading2"],
        )
    )

    # Orden eliminada y nombres agregados.
    detail_columns = [
        "Archivo",
        "Equipo",
        "Técnico responsable",
        "Jefe de turno",
        "Sección",
        "Campo faltante",
        "Prioridad",
        "Celda/Rango",
    ]

    detail_show = (
        detail_df[detail_columns].head(30).copy()
        if not detail_df.empty
        else pd.DataFrame(columns=detail_columns)
    )

    # Mostrar texto informativo cuando el nombre está vacío.
    for column in ["Técnico responsable", "Jefe de turno"]:
        detail_show[column] = (
            detail_show[column]
            .fillna("")
            .replace("", "Sin información")
        )

    detail_table_data = (
        [detail_show.columns.tolist()]
        + detail_show.astype(str).values.tolist()
    )

    detail_table = Table(
        detail_table_data,
        repeatRows=1,
        colWidths=[
            5.2 * cm,
            1.7 * cm,
            3.2 * cm,
            3.2 * cm,
            3.0 * cm,
            3.0 * cm,
            1.8 * cm,
            2.0 * cm,
        ],
    )

    detail_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#FFC400"),
                ),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    story.append(detail_table)

    document.build(story)
    return output.getvalue()


def render_validation_app() -> None:
    # ============================================================
    # Sidebar
    # ============================================================
    with st.sidebar:
        # Acceso nativo al Centro de Aplicaciones.
        # Se ubica sobre el logo y abre el centro en una pestaña nueva.
        st.link_button(
            "CENTRO DE APLICACIONES",
            URL_CENTRO_APLICACIONES,
            use_container_width=True,
        )

        if LOGO_PATH.exists():
            st.image(
                str(LOGO_PATH),
                use_container_width=True,
            )
        else:
            st.warning(
                "No se encontró el archivo Finning-CAT.png "
                "en la carpeta del proyecto."
            )

        st.markdown("---")

        st.markdown(
            (
                '<div class="step-card">'
                '<span class="step-number">1</span>'
                '<b>Subir archivos Excel</b><br>'
                '<small>Formatos: .xlsx, .xlsm o .zip con varias OT.</small>'
                "</div>"
            ),
            unsafe_allow_html=True,
        )

        uploaded_files = st.file_uploader(
            "Arrastra y suelta tus órdenes aquí",
            type=["xlsx", "xlsm", "zip"],
            accept_multiple_files=True,
            label_visibility="collapsed",
        )

        st.markdown(
            (
                '<div class="step-card">'
                '<span class="step-number">2</span>'
                '<b>Validar órdenes</b><br>'
                '<small>Se revisan únicamente los campos esenciales definidos.</small>'
                "</div>"
            ),
            unsafe_allow_html=True,
        )

        validate_btn = st.button(
            "Validar órdenes",
            use_container_width=True,
        )

        st.markdown(
            (
                '<div class="step-card">'
                '<span class="step-number">3</span>'
                '<b>Descargar reporte</b><br>'
                '<small>Se generan reportes en Excel y PDF.</small>'
                "</div>"
            ),
            unsafe_allow_html=True,
        )


    # ============================================================
    # Pantalla principal
    # ============================================================
    st.markdown(
        '<p class="main-title">REPORTE EJECUTIVO DE VALIDACIÓN</p>',
        unsafe_allow_html=True,
    )

    st.markdown(
        (
            '<p class="subtitle">'
            "Validación de horómetro, motivo de detención, síntoma, "
            "códigos principales, descripción de actividades y firmas "
            "de la Orden de Trabajo."
            "</p>"
        ),
        unsafe_allow_html=True,
    )


    if "results" not in st.session_state:
        st.session_state.results = None


    if validate_btn:
        if not uploaded_files:
            st.warning(
                "Debes subir al menos un archivo Excel o ZIP para validar."
            )
        else:
            files = read_uploaded_files(uploaded_files)

            if not files:
                st.error(
                    "No se encontraron archivos .xlsx o .xlsm válidos."
                )
            else:
                progress = st.progress(
                    0,
                    text="Validando órdenes...",
                )

                results: List[Dict[str, Any]] = []
                errors: List[Dict[str, str]] = []

                for index, (filename, data) in enumerate(
                    files,
                    start=1,
                ):
                    try:
                        results.append(
                            validate_work_order(
                                data,
                                filename,
                            )
                        )
                    except Exception as error:
                        errors.append(
                            {
                                "Archivo": filename,
                                "Error": str(error),
                            }
                        )

                    progress.progress(
                        index / len(files),
                        text=(
                            f"Validando {index} de {len(files)}: "
                            f"{filename}"
                        ),
                    )

                progress.empty()

                st.session_state.results = {
                    "results": results,
                    "errors": errors,
                }


    if st.session_state.results is None:
        st.info(
            "Sube una o varias órdenes de trabajo y presiona "
            "Validar órdenes para generar el reporte."
        )
        st.stop()


    results = st.session_state.results["results"]
    errors = st.session_state.results["errors"]


    if not results:
        st.error(
            "No fue posible procesar ninguna orden de trabajo."
        )

        if errors:
            st.dataframe(
                pd.DataFrame(errors),
                use_container_width=True,
                hide_index=True,
            )

        st.stop()


    summary_df, detail_df = create_result_dataframes(results)

    total_orders = len(summary_df)

    field_summary_df = build_field_summary(
        detail_df,
        total_orders,
    )

    compliance_summary_df = build_compliance_summary(
        detail_df,
        total_orders,
    )

    section_summary_df = build_section_summary(detail_df)

    orders_with_obs = int(
        (
            summary_df["Estado"]
            == "Con observaciones"
        ).sum()
    )

    complete_orders = int(
        (
            summary_df["Estado"]
            == "Completa"
        ).sum()
    )

    total_missing = int(
        summary_df["Campos faltantes"].sum()
    )

    critical_missing = int(
        summary_df["Faltantes críticos"].sum()
    )


    # ============================================================
    # KPIs
    # ============================================================
    kpi_columns = st.columns(5)

    kpis = [
        (
            "Órdenes procesadas",
            total_orders,
            "Total de OT procesadas",
        ),
        (
            "Órdenes con observaciones",
            orders_with_obs,
            "OT con al menos un hallazgo",
        ),
        (
            "Hallazgos detectados",
            total_missing,
            "Suma de todos los hallazgos detectados",
        ),
        (
            "Hallazgos críticos",
            critical_missing,
            "Códigos o firmas",
        ),
        (
            "Órdenes completas",
            complete_orders,
            "Sin hallazgos detectados",
        ),
    ]

    for column, (title, value, note) in zip(
        kpi_columns,
        kpis,
    ):
        with column:
            st.markdown(
                (
                    '<div class="kpi-card">'
                    f'<div class="kpi-title">{title}</div>'
                    f'<div class="kpi-value">{value}</div>'
                    f'<div class="kpi-note">{note}</div>'
                    "</div>"
                ),
                unsafe_allow_html=True,
            )


    st.write("")


    # ============================================================
    # Gráficos
    # ============================================================
    left_chart, right_chart = st.columns([1.45, 1])


    with left_chart:
        st.markdown(
            '<div class="panel">',
            unsafe_allow_html=True,
        )

        st.subheader("Cumplimiento por campo")

        st.caption(
            "Cada barra representa el 100% de las OT procesadas "
            "y facilita comparar todos los campos."
        )

        chart_df = compliance_summary_df.melt(
            id_vars=["Campo"],
            value_vars=[
                "% Cumplimiento",
                "% Con observación",
            ],
            var_name="Resultado",
            value_name="Porcentaje",
        )

        chart_df["Resultado"] = chart_df["Resultado"].replace(
            {
                "% Cumplimiento": "Conforme",
                "% Con observación": "Con observación",
            }
        )

        chart_df["Etiqueta"] = chart_df[
            "Porcentaje"
        ].apply(
            lambda value: (
                f"{value:.1f}%"
                if value >= 4
                else ""
            )
        )

        figure = px.bar(
            chart_df,
            x="Porcentaje",
            y="Campo",
            orientation="h",
            color="Resultado",
            text="Etiqueta",
            barmode="stack",
            color_discrete_map={
                "Conforme": OK_COLOR,
                "Con observación": OBS_COLOR,
            },
            category_orders={
                "Campo": list(reversed(VALIDATED_FIELDS)),
                "Resultado": [
                    "Conforme",
                    "Con observación",
                ],
            },
            template="plotly_white",
        )

        figure.update_traces(
            textposition="inside",
            insidetextanchor="middle",
        )

        figure.update_layout(
            height=max(
                430,
                42 * len(VALIDATED_FIELDS),
            ),
            margin=dict(
                l=10,
                r=20,
                t=10,
                b=10,
            ),
            xaxis_title="Porcentaje de OT",
            yaxis_title="",
            xaxis=dict(
                range=[0, 100],
                ticksuffix="%",
            ),
            legend_title_text="Resultado",
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(color="#111111"),
        )

        st.plotly_chart(
            figure,
            use_container_width=True,
        )

        st.markdown(
            "</div>",
            unsafe_allow_html=True,
        )


    with right_chart:
        st.markdown(
            '<div class="panel">',
            unsafe_allow_html=True,
        )

        st.subheader("Estado de órdenes")

        status_df = pd.DataFrame(
            {
                "Estado": [
                    "Completas",
                    "Con observaciones",
                ],
                "Cantidad": [
                    complete_orders,
                    orders_with_obs,
                ],
            }
        )

        figure = px.pie(
            status_df,
            values="Cantidad",
            names="Estado",
            hole=0.55,
            color="Estado",
            color_discrete_map={
                "Completas": OK_COLOR,
                "Con observaciones": OBS_COLOR,
            },
            template="plotly_white",
        )

        figure.update_layout(
            margin=dict(
                l=10,
                r=10,
                t=10,
                b=10,
            ),
            height=360,
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(color="#111111"),
        )

        st.plotly_chart(
            figure,
            use_container_width=True,
        )

        completion_percentage = (
            complete_orders
            / total_orders
            * 100
            if total_orders
            else 0
        )

        st.info(
            f"{complete_orders} de {total_orders} órdenes "
            f"({completion_percentage:.1f}%) están completas."
        )

        st.markdown(
            "</div>",
            unsafe_allow_html=True,
        )


    st.write("")


    # ============================================================
    # Resúmenes
    # ============================================================
    summary_tab, field_tab, detail_tab = st.tabs(
        [
            "Resumen por OT",
            "Resumen de campos faltantes",
            "Detalle de hallazgos",
        ]
    )


    with summary_tab:
        st.dataframe(
            summary_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Campos con observación": st.column_config.TextColumn(
                    "Campos con observación",
                    help=(
                        "Identifica los campos faltantes o inválidos "
                        "detectados en cada OT."
                    ),
                    width="large",
                )
            },
        )


    with field_tab:
        if field_summary_df.empty:
            st.success(
                "No existen campos faltantes para resumir."
            )
        else:
            st.dataframe(
                field_summary_df,
                use_container_width=True,
                hide_index=True,
            )

            st.markdown(
                "#### Campos faltantes por sección"
            )

            st.dataframe(
                section_summary_df,
                use_container_width=True,
                hide_index=True,
            )


    with detail_tab:
        if detail_df.empty:
            st.success(
                "Todas las órdenes revisadas están completas."
            )
        else:
            # La columna Orden fue eliminada.
            # Se agregan Técnico responsable y Jefe de turno.
            filtered_detail = detail_df[
                [
                    "Archivo",
                    "Técnico responsable",
                    "Jefe de turno",
                    "Campo faltante",
                    "Observación",
                ]
            ].copy()

            filtered_detail = filtered_detail.rename(
                columns={
                    "Archivo": "Archivo OT",
                    "Campo faltante": "Hallazgo",
                    "Observación": "Detalle",
                }
            )

            for column in [
                "Técnico responsable",
                "Jefe de turno",
            ]:
                filtered_detail[column] = (
                    filtered_detail[column]
                    .fillna("")
                    .replace("", "Sin información")
                )

            st.dataframe(
                filtered_detail,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Archivo OT": st.column_config.TextColumn(
                        "Archivo OT",
                        width="large",
                    ),
                    "Técnico responsable": st.column_config.TextColumn(
                        "Técnico responsable",
                        help=(
                            "Nombre registrado en el recuadro de firma "
                            "del técnico responsable."
                        ),
                        width="medium",
                    ),
                    "Jefe de turno": st.column_config.TextColumn(
                        "Jefe de turno",
                        help=(
                            "Nombre registrado en el recuadro de firma "
                            "del jefe de turno."
                        ),
                        width="medium",
                    ),
                    "Hallazgo": st.column_config.TextColumn(
                        "Hallazgo",
                        width="medium",
                    ),
                    "Detalle": st.column_config.TextColumn(
                        "Detalle",
                        width="large",
                    ),
                },
            )


    if errors:
        with st.expander("Archivos no procesados"):
            st.dataframe(
                pd.DataFrame(errors),
                use_container_width=True,
                hide_index=True,
            )


    # ============================================================
    # Descargas
    # ============================================================
    excel_bytes = build_excel_report(
        summary_df,
        detail_df,
        field_summary_df,
        section_summary_df,
    )

    pdf_bytes = build_pdf_report(
        summary_df,
        detail_df,
        field_summary_df,
    )

    download_excel, download_pdf, _ = st.columns(
        [1, 1, 2]
    )


    with download_excel:
        st.download_button(
            "Descargar Excel",
            data=excel_bytes,
            file_name=(
                "reporte_validacion_ot_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
        )


    with download_pdf:
        st.download_button(
            "Descargar PDF",
            data=pdf_bytes,
            file_name=(
                "reporte_ejecutivo_ot_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            ),
            mime="application/pdf",
            use_container_width=True,
        )


    with st.expander("Campos revisados en el formato OT"):
        st.markdown(
            """
            La validación considera exclusivamente los siguientes campos:

            - *Horómetro:* debe contener un valor distinto de *0*.
            - *Motivo de detención del equipo*.
            - *Descripción del síntoma*.
            - *Código trabajo*.
            - *Código síntoma*.
            - *Código causa:* los valores *6.6* y *7.1* se consideran inválidos porque corresponden a la categoría “Otros”.
            - *Descripción de actividades*.
            - *Firma del jefe de turno*, verificando nombre y RUT.
            - *Firma del técnico responsable*, verificando nombre y RUT.

            No se generan observaciones por ningún otro campo del anverso o reverso.
            """
        )

def get_supabase_repository() -> Optional[SupabaseRepository]:
    if create_client is None:
        st.error(
            "No se encuentra instalada la librería de Supabase. "
            "Revisa requirements.txt."
        )
        return None

    try:
        config = st.secrets["supabase"]

        url = str(config.get("url", "")).strip().rstrip("/")
        key = str(config.get("secret_key", "")).strip()

        if not url:
            st.error("Falta configurar supabase.url en Streamlit Secrets.")
            return None

        if not key:
            st.error(
                "Falta configurar supabase.secret_key "
                "en Streamlit Secrets."
            )
            return None

        if not key.startswith("sb_secret_"):
            st.error(
                "La clave configurada no es una Secret key de Supabase. "
                "Debe comenzar con 'sb_secret_'. No utilices "
                "'sb_publishable_' ni la clave anon."
            )
            return None

        client = create_client(url, key)
        return SupabaseRepository(client)

    except KeyError:
        st.error(
            "No existe la sección [supabase] en Streamlit Secrets."
        )
        return None

    except Exception as exc:
        st.error(f"No fue posible conectar con Supabase: {exc}")
        return None


def render_weekly_control() -> None:
    repo = get_supabase_repository()

    with st.sidebar:
        st.link_button(
            "CENTRO DE APLICACIONES",
            URL_CENTRO_APLICACIONES,
            use_container_width=True,
        )
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), use_container_width=True)
        st.markdown("---")
        st.markdown(
            '<div class="step-card"><span class="step-number">1</span>'
            '<b>Cargar detenciones</b><br><small>Reporte Excel con hoja Limpio.</small></div>',
            unsafe_allow_html=True,
        )
        detention_file = st.file_uploader(
            "Reporte de detenciones",
            type=["xlsx", "xlsm"],
            key="weekly_detentions_uploader",
            label_visibility="collapsed",
        )
        exclusions_text = st.text_area(
            "Categorías excluidas",
            value="\n".join(DEFAULT_EXCLUSIONS),
            help="Una palabra o categoría por línea. Estas detenciones no exigirán OT.",
        )
        import_detentions = st.button(
            "Importar detenciones",
            use_container_width=True,
            key="import_weekly_detentions",
        )
        st.markdown(
            '<div class="step-card"><span class="step-number">2</span>'
            '<b>Guardar OT digitales</b><br><small>Usa las OT procesadas en Validación OT.</small></div>',
            unsafe_allow_html=True,
        )
        has_validated = bool(st.session_state.get("results") and st.session_state.results.get("results"))
        save_ots = st.button(
            "Guardar OT validadas",
            use_container_width=True,
            disabled=not has_validated,
            key="save_weekly_ots",
        )

    st.markdown('<p class="main-title">CONTROL DE OT SEMANAL</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="subtitle">Seguimiento diario de detenciones respaldadas con órdenes de trabajo digitales. '
        'Las OT recibidas posteriormente se asignan al día original de la detención.</p>',
        unsafe_allow_html=True,
    )

    if repo is None:
        st.warning(
            "Supabase aún no está configurado. La aplicación de validación seguirá funcionando, "
            "pero el Control OT Semanal necesita las credenciales en Streamlit Secrets."
        )
        schema_path = APP_DIR / "supabase_schema.sql"
        if schema_path.exists():
            st.download_button(
                "Descargar estructura SQL para Supabase",
                data=schema_path.read_bytes(),
                file_name="supabase_schema.sql",
                mime="text/plain",
            )
        st.code(
            '[supabase]\nurl = "https://TU-PROYECTO.supabase.co"\n'
            'service_role_key = "TU_SERVICE_ROLE_KEY"',
            language="toml",
        )
        st.info(
            "Crea el proyecto en Supabase, ejecuta el archivo SQL en SQL Editor y agrega estas "
            "credenciales en Settings > Secrets de Streamlit Cloud."
        )
        return

    if import_detentions:
        if detention_file is None:
            st.warning("Selecciona primero el reporte de detenciones.")
        else:
            try:
                exclusions = [line.strip() for line in exclusions_text.splitlines() if line.strip()]
                det_df = read_detentions_excel(
                    detention_file.getvalue(), detention_file.name, exclusions
                )
                imported = repo.upsert_detentions(det_df.to_dict("records"))
                st.success(f"Se importaron o actualizaron {imported} detenciones.")
            except Exception as error:
                st.error(f"No fue posible importar el reporte: {error}")

    if save_ots:
        try:
            validation_results = st.session_state.results["results"]
            records = work_order_records(validation_results)
            imported = repo.upsert_work_orders(records)
            reference = date.today()
            start, end = week_bounds(reference)
            associated = repo.auto_associate(start - timedelta(days=30), end)
            st.success(
                f"Se guardaron o actualizaron {imported} OT digitales. "
                f"Asociaciones automáticas realizadas: {associated}."
            )
        except Exception as error:
            st.error(f"No fue posible guardar las OT: {error}")

    filter_col1, filter_col2, filter_col3 = st.columns([1, 1, 1.3])
    today = date.today()
    default_start, default_end = week_bounds(today)
    with filter_col1:
        start_date = st.date_input("Desde", value=default_start, key="weekly_start")
    with filter_col2:
        end_date = st.date_input("Hasta", value=default_end, key="weekly_end")
    with filter_col3:
        st.write("")
        st.write("")
        refresh = st.button("Actualizar datos", use_container_width=True, key="refresh_weekly")

    try:
        detentions = repo.list_detentions(start_date, end_date)
        detention_ids = detentions["id"].tolist() if not detentions.empty else []
        associations = repo.list_associations(detention_ids)
        ots = repo.list_work_orders()
    except Exception as error:
        st.error(f"No fue posible consultar Supabase: {error}")
        return

    valid = detentions[
        detentions.get("requiere_ot", pd.Series(dtype=bool)).fillna(False).astype(bool)
    ].copy() if not detentions.empty else pd.DataFrame()
    associated_detention_ids = set(
        associations.get("detencion_id", pd.Series(dtype=object)).dropna().astype(str)
    ) if not associations.empty else set()
    if not valid.empty:
        valid["con_ot"] = valid["id"].astype(str).isin(associated_detention_ids)
    total_valid = len(valid)
    backed = int(valid["con_ot"].sum()) if not valid.empty else 0
    pending_count = total_valid - backed
    compliance = backed / total_valid * 100 if total_valid else 0.0

    kpi_columns = st.columns(4)
    kpis = [
        ("Detenciones válidas", total_valid, "Total del período"),
        ("Detenciones con OT", backed, "OT digital asociada"),
        ("Pendientes", pending_count, "Sin respaldo digital"),
        ("Cumplimiento actualizado", f"{compliance:.1f}%", "Asignado al día de detención"),
    ]
    for column, (title, value, note) in zip(kpi_columns, kpis):
        with column:
            st.markdown(
                '<div class="kpi-card">'
                f'<div class="kpi-title">{title}</div>'
                f'<div class="kpi-value">{value}</div>'
                f'<div class="kpi-note">{note}</div></div>',
                unsafe_allow_html=True,
            )

    st.write("")
    st.subheader("Resumen diario")
    daily_df = daily_summary(detentions, associations)
    if daily_df.empty:
        st.info("No existen detenciones válidas para el período seleccionado.")
    else:
        st.dataframe(
            daily_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Cumplimiento (%)": st.column_config.ProgressColumn(
                    "Cumplimiento (%)", min_value=0, max_value=100, format="%.1f%%"
                )
            },
        )

    st.write("")
    st.subheader("Detalle de OT pendientes")
    if valid.empty:
        pending = pd.DataFrame()
    else:
        pending = valid[~valid["con_ot"]].copy()

    if pending.empty:
        st.success("No existen detenciones pendientes de OT en el período seleccionado.")
    else:
        pending_display = pd.DataFrame({
            "Fecha": pd.to_datetime(pending["fecha_detencion"], errors="coerce").dt.strftime("%d/%m/%Y"),
            "Hora": pending.get("hora_inicio", "").fillna("").astype(str).str.slice(0, 5),
            "Turno": pending.get("turno", "").fillna(""),
            "Equipo": pending.get("equipo", "").fillna(""),
            "Descripción / Razón": pending.get("razon", "").fillna(""),
            "Comentario": pending.get("comentario", "").fillna(""),
        })
        st.dataframe(pending_display, use_container_width=True, hide_index=True)

        with st.expander("Asociar manualmente una OT pendiente"):
            detention_options = {
                f"{row.get('fecha_detencion', '')} {str(row.get('hora_inicio', ''))[:5]} | "
                f"{row.get('equipo', '')} | {row.get('razon', '')}": row.get("id")
                for _, row in pending.iterrows()
            }
            selected_detention_label = st.selectbox(
                "Detención", list(detention_options.keys()), key="manual_detention"
            )
            selected_detention_id = detention_options[selected_detention_label]
            selected_detention = pending[pending["id"] == selected_detention_id].iloc[0].to_dict()

            available_ots = ots.copy()
            if not available_ots.empty:
                available_ots = available_ots[
                    available_ots["equipo"].apply(normalize_weekly_text)
                    == normalize_weekly_text(selected_detention.get("equipo"))
                ]
            if available_ots.empty:
                st.info("No hay OT digitales guardadas para este equipo.")
            else:
                available_ots["score"] = available_ots.apply(
                    lambda row: similarity(selected_detention, row.to_dict()), axis=1
                )
                available_ots = available_ots.sort_values("score", ascending=False)
                ot_options = {
                    f"OT {row.get('numero_ot', '')} | Coincidencia {row.get('score', 0)*100:.0f}% | "
                    f"{row.get('descripcion', '')[:90]}": row.get("id")
                    for _, row in available_ots.iterrows()
                }
                selected_ot_label = st.selectbox("OT digital", list(ot_options.keys()), key="manual_ot")
                selected_ot_id = ot_options[selected_ot_label]
                selected_score = float(
                    available_ots.loc[available_ots["id"] == selected_ot_id, "score"].iloc[0]
                )
                if st.button("Confirmar asociación", key="confirm_manual_association"):
                    try:
                        repo.associate(selected_detention_id, selected_ot_id, selected_score, "MANUAL")
                        st.success("OT asociada correctamente al día original de la detención.")
                        st.rerun()
                    except Exception as error:
                        st.error(f"No fue posible asociar la OT: {error}")

    st.caption(
        "Una detención puede tener una o varias OT asociadas. Para el cumplimiento diario, "
        "la detención se considera respaldada cuando posee al menos una OT digital confirmada."
    )


st.markdown(
    """
    <style>
    div[role="radiogroup"] {gap: 0.35rem; margin-bottom: 1rem;}
    div[role="radiogroup"] label {
        background: #ffffff; border: 1px solid #dfe3e8; border-radius: 8px;
        padding: 0.45rem 0.9rem; font-weight: 800;
    }
    div[role="radiogroup"] label:has(input:checked) {
        background: #ffc400; border-color: #ffc400; color: #111111;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

module = st.radio(
    "Módulo",
    ["Validación OT", "Control OT Semanal"],
    horizontal=True,
    label_visibility="collapsed",
)

if module == "Validación OT":
    render_validation_app()
else:
    render_weekly_control()

